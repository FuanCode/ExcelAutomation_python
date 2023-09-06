import pandas as pd
from openpyxl import load_workbook

# Nome dos arquivos Excel
excel1_file = 'local do arquivo'
excel2_file = 'local do arquivo'

# Carregando os dados do Excel 1 em um DataFrame
df_excel1 = pd.read_excel(excel1_file)

# Escolha a linha e coluna onde deseja colar o conteúdo do Excel 1 no Excel 2
linha_destino = 2  # Exemplo: linha 2
coluna_destino = 'A'  # Exemplo: coluna C

# Carregando o arquivo Excel 2 usando openpyxl
excel2 = load_workbook(excel2_file)
writer = pd.ExcelWriter(excel2_file, engine='openpyxl')
writer.book = excel2

# Copiando o conteúdo do DataFrame do Excel 1 para o Excel 2
df_excel1.to_excel(writer, sheet_name='Planilha1', startrow=linha_destino - 1, startcol=writer.sheets['Planilha1'].max_column + 1, index=False, header=False)

# Salvar as mudanças no arquivo Excel 2
writer.save()

print(f"Conteúdo do {excel1_file} copiado com sucesso para {excel2_file} na linha {linha_destino} e coluna {coluna_destino}.")


#Neste código:

#pd.read_excel() é usado para carregar os dados do Excel 1 em um DataFrame Pandas.
#Você pode definir linha_destino e coluna_destino para escolher onde deseja colar o conteúdo do Excel 1 no Excel 2.
#load_workbook é usado para carregar o arquivo Excel 2 com a biblioteca openpyxl.
#to_excel() é usado para copiar o conteúdo do DataFrame do Excel 1 para o Excel 2 na posição especificada.
#Por fim, save() é usado para salvar as mudanças no arquivo Excel 2.
#Lembre-se de substituir 'excel1.xlsx' e 'excel2.xlsx' pelos nomes dos seus próprios arquivos Excel e 
#ajustar linha_destino e coluna_destino conforme necessário. Certifique-se de que as bibliotecas Pandas e openpyxl estejam instaladas no seu ambiente Python.